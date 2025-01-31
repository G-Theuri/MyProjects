import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import time, os, random, requests
from rich import print
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

def get_data(driver, url, image_path):
    driver.execute_script(f'window.open("{url}", "_blank");')
    driver.switch_to.window(driver.window_handles[1])
    time.sleep(3)
    
    #close the location enquiries dialog box
    try:
        WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//button[@class="ot-close-icon"]'))).click()
    except:
        pass
    


    #Get products data
    try:
        title = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//div[@class="pdp-title-section v2"]/h1'))).text
    except:
        title =''
    try:
        image = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//button[@class="Ub-Mh_gf"]/img'))).get_attribute('src')

    except:
        image =''
    try:
        price = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//div[@class="sale-subscription-price-block"]/span'))).text
    except:
        price =''
    try:
        description = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//section[@id="overview"]/div/div/div/div'))).text
    except:
        description =''


    all_specs = {}
    try:
        specs = driver.find_elements(By.XPATH, '//*[@id="detailedSpecs"]/div/div/div/div')
        for spec in specs:
            label = spec.find_element(By.XPATH, './/div[@class ="Ea-Ee_gf"]/p').text
            value = spec.find_element(By.XPATH, './/p[@class ="Cv-B_gf Cv-C7_gf Ea-Eg_gf Cv-K_gf"]/span').text
            all_specs[label] = value
    except:
        pass

    if image:
        try:
            response = requests.get(image)
            if response.status_code == 200:
                with open(image_path, 'wb') as file:
                    file.write(response.content)

            
        except Exception as e:
            print(f"Error downloading image: {e}")
            
    #load acquired data into a dictionary
    data ={
            'Title': title,
            'URL': url,
            'Image': image, 
            'Price': price,
            'Description': description,
            'Specs': all_specs
        }

    return data

def add_image_to_excel(image_path, wb, row, column_name):
    
    sheet = wb.active

    #column_index =sheet.columns.index(sheet[column_name]) + 1

    img = Image(image_path)
    sheet.add_image(img, f'{column_name}{row}')

    img_width, img_height = img.width, img.height
    scale_factor = 0.75
    sheet.row_dimensions[row].height = img_height * scale_factor

    column_width = img_width / 7
    sheet.column_dimensions['AN'].width = column_width

def main():
    filepath = 'C:/Users/TG/Downloads/WebScrape-Content-Template.xlsx'

    if not os.path.exists('images'):
        os.makedirs('images')
    
    options =uc.ChromeOptions()
    options.add_argument('--disable-popup-blocking')
    driver =uc.Chrome(options)
    driver.maximize_window()

    max_retries = 2
    retry_delay = 1

    excel_filename = 'new - updated_file.xlsx'
    if not os.path.exists(excel_filename):
        wb = Workbook() # Create a new workbook if it doesn't exist
        wb.save(excel_filename)
    else:
        wb = load_workbook(excel_filename) # Load the existing workbook

    try:
        driver.get('https://www.hp.com/us-en/home.html')
        try:
            WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="onetrust-close-btn-container"]/button'))).click()
        except:
            pass

        searchbar = driver.find_element(By.XPATH, '//*[@id="search_focus_desktop"]')
        searchbar.clear()

        df = pd.read_excel(filepath, sheet_name='HP')

        for index, row in df.iterrows():
            model_name = row['model name']
            model_number = row['mfr number']
            image_path = f'images/{model_number}.jpg'

            searchbar.clear()
            searchbar.send_keys(model_number)
            time.sleep(4)
            
            retries = 0
            success = False

            while retries < max_retries and not success:
                try:
                    
                    #Check if searching using model_number yields any suggestion and if not, use the model_name
                    try:
                        suggestion = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//div[@class="shop ac-cards"]/a')))
                    except:
                        searchbar.clear()
                        searchbar.send_keys(model_name)
                        time.sleep(2)
                        suggestion = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//div[@class="shop ac-cards"]/a')))

                    item_url = suggestion.get_attribute('href')
                    data = get_data(driver, item_url, image_path)

                    if data:
                        print(f"[yellow]{model_number}[/yellow]: {data}")

                        #Resize image
                        image = PILImage.open(image_path)
                        if image.mode in ("RGBA", "LA"):
                            image = image.convert("RGB")
                        image = image.resize((100, 100), PILImage.Resampling.LANCZOS) 
                        image.save(image_path)

                        #Load the acquired data
                        df.at[index, 'Product URL'] = data['URL']
                        df.at[index, 'unit cost'] = data['Price'].replace('$', '')
                        df.at[index, 'Product Image'] = data['Image']
                        df.at[index, 'product description'] = data['Description']
                        
                        try:
                            df.at[index, 'weight'] = data['Specs']['Weight']
                        except:
                            pass

                        try:
                            dimension= data['Specs']['Dimensions (W X D X H)'].split(' x ')
        
                            df.at[index, 'depth'] = dimension[1]
                            df.at[index, 'height'] = dimension[-1].replace(' in', '')
                            df.at[index, 'width'] = dimension[0]
                        except:
                            pass
                        
                        # Save the updated data (only data, not images yet)
                        df.to_excel(excel_filename, index=False, sheet_name='HP')
                        if data['Image'] != '':
                            # Add image to Excel file
                            add_image_to_excel(image_path, wb, row=index +2, column_name='AN')
                        
                        success = True
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])

                    else:
                        print(f"No data returned for {model_number}")
                        break # Exit retry loop if no data is returned

                except (NoSuchElementException, TimeoutException) as e:
                    retries += 1

                    if retries < max_retries:
                        time.sleep(retry_delay)
                    else:
                        print(f'[yellow]{model_number}[/yellow] [red]Not found![/red]')
                        break # Exit loop after retries are exhausted

                except Exception as e:
                    print(f"[yellow]{model_number}[/yellow] [red]Unexpected error: {str(e)}[/red]")
                    break  # Exit loop for any other unexpected errors
        
        # After all images are added, save the workbook
        wb.save(excel_filename)

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        driver.quit()

if __name__ == "__main__":
    main()