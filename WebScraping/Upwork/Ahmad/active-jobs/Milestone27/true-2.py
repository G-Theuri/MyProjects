import camelot
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
import re

# Read the PDF using Camelot (True Spec Sheet PDF)
pdf_path = "resources/TRUE SPEC SHEET.pdf"
tables = camelot.read_pdf(pdf_path, pages='1', flavor='lattice')

### Extract the Specs Table
df_specs = tables[3].df
df_specs_clean = df_specs.dropna(how='all', axis=0).dropna(how='all', axis=1)
df_specs_clean = df_specs_clean.drop(columns=[0, 1]).drop(index=range(11, 14)).drop(index=range(0, 2)).reset_index(drop=True)
split_data = df_specs_clean[2].str.split('\n', expand=True)
split_data.columns = ['Dimensions', 'in.', 'mm']

# Split the Specs Table into Dimensions and Electrical
df_dimensions = split_data.iloc[0:3].reset_index(drop=True)
df_electrical = split_data.iloc[4:].reset_index(drop=True)
df_electrical.columns = ['Electrical', 'U.S.', 'International']
df_electrical.drop(0).reset_index(drop=True)

### Extract models data and add Dimensions and Electrical data
models_data = []
for t in range(0, 3):
    df = tables[t].df
    df_clean = df.dropna(how='all', axis=0).dropna(how='all', axis=1)
    df_clean = df_clean.drop(columns=[0]).reset_index(drop=True) if t == 0 else df_clean

    model_name = df_clean.iloc[0, 0]
    description = {}
    
    # Extracting description data for each model
    for i in range(1, len(df_clean)):
        key = df_clean.iloc[i, 0]
        value = df_clean.iloc[i, 1].replace(" \n", '')
        description[key] = value

    # Extract Width, Depth, Height from the dimensions table
    width = df_dimensions.loc[0, 'in.']
    depth = df_dimensions.loc[1, 'in.']
    height = df_dimensions.loc[2, 'in.']

    # Store each model's data including width, depth, and height as separate attributes
    models_data.append({
        'Model': model_name,
        'Width': width,
        'Depth': depth,
        'Height': height,
        'HP': df_electrical.loc[0, 'International'],
        'Amps': df_electrical.loc[1, 'U.S.'],
        'Voltage': df_electrical.loc[2, 'U.S.'],
        'NEMA Connection': df_electrical.loc[3, 'U.S.'],
        'Description': description,
    })

# Create a new workbook using openpyxl
wb = Workbook()

# Define the color for section separator rows
separator_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
model_fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")

# Remove default sheet and add data for each model
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Iterate over each model and create a new sheet for each one
for entry in models_data:
    model_name = entry.get('Model', 'Unknown Model')
    sheet = wb.create_sheet(title=model_name)
    
    # Set column headers: Attributes, Values, Notes
    rw = 1
    sheet.cell(rw, column=1, value="Attributes").alignment = Alignment(wrap_text=True)
    sheet.cell(rw, column=2, value="Values").alignment = Alignment(wrap_text=True)
    sheet.cell(rw, column=3, value="Notes").alignment = Alignment(wrap_text=True)
    for col in range(3):
        sheet[rw][col].fill = header_fill

    # Define the sections (Attributes) and Values
    sections = [
        ("Model", entry.get('Model', '')),
        ("External Dimensions", ""), ("Width", entry.get('Width', '')), ("Depth", entry.get('Depth', '')), ("Height", entry.get('Height', '')),
        ("Shipping Info", ""), ("Certifications", ""),
        ("Electrical", ""), ("Voltage", entry.get('Voltage', '')), ("Cycle", ""), ("Phase", ""), ("Amps", entry.get('Amps', '')), ("Kw", entry.get('HP', '')), ("HP", ""), ("NEMA Connection", entry.get('NEMA Connection', '')),
        ("Gas", ""), ("Gas Type (Natural or Propane)", ""), ("Gas Size", ""), ("Gas Total BTUs", ""), ("Gas Kw", ""),
        ("Plumbing", ""), ("Cold Water Size", ""), ("Hot Water Size", ""), ("Drain Size", ""), ("Indirect Waste Size", ""), ("Direct Waste Size", "")
    ]
    
    # Write sections to the sheet
    row = 2
    for section in sections:
        # Write the attribute (e.g., "Model", "Voltage", etc.)
        sheet.cell(row=row, column=1, value=section[0]).alignment = Alignment(wrap_text=True)
        sheet.cell(row=row, column=2, value=section[1]).alignment = Alignment(wrap_text=True)
        sheet.cell(row=row, column=3, value="").alignment = Alignment(wrap_text=True)

        # Apply the color to separator rows
        if section[0] == "Model":
            for col in range(3):
                sheet[row][col].fill = model_fill

        if section[0] == "External Dimensions" or section[0] == "Shipping Info" or section[0] == "Electrical" or section[0] == "Gas" or section[0] == "Plumbing":
            for col in range(3):
                sheet[row][col].fill = separator_fill
        row += 1

# Save the workbook to a file
wb.save('TRUE-SPEC-SHEET-WORKBOOK.xlsx')
print("Excel workbook 'TRUE-SPEC-SHEET-WORKBOOK.xlsx' has been created.")
