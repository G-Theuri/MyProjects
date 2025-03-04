import camelot
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
import re

# Read the PDF using Camelot
pdf_path = "resources/Hatco Spec Sheet.pdf"
tables = camelot.read_pdf(pdf_path, pages='2', flavor='stream')

# Extract the data from the first table
df = tables[0].df
df_clean = df.dropna(how='all', axis=0).dropna(how='all', axis=1)
df_clean = df_clean.iloc[4:]  # Clean up and ignore first 4 rows (headers)

# Initialize a list to store the data points
data_points = []
current_data = {}

# Iterate through the cleaned-up DataFrame and process each row
for index, row in df_clean.iterrows():
    if 'NEMA 5-15P' in str(row[7]) and current_data:
        data_points.append(current_data)
        current_data = {}  # Reset current_data for the next model

    if pd.notna(row[0]):  # Model column
        current_data['Model'] = current_data.get('Model', '') + " " + str(row[0]).strip()

    if pd.notna(row[2]):  # Dimensions column
        current_data['External Dimensions'] = current_data.get('External Dimensions', '') + " " + str(row[2]).strip()

    if pd.notna(row[3]):  # Voltage column
        current_data['Voltage'] = current_data.get('Voltage', '') + " " + str(row[3]).strip()

    if pd.notna(row[4]):  # Watts column
        current_data['Watts'] = current_data.get('Watts', '') + " " + str(row[4]).strip()

    if pd.notna(row[6]):  # Amps column
        current_data['Amps'] = current_data.get('Amps', '') + " " + str(row[6]).strip()

    if pd.notna(row[8]):  # Shipping Weight column
        current_data['Shipping Weight'] = current_data.get('Shipping Weight', '') + " " + str(row[8]).strip()

    if pd.notna(row[7]):  # Plugs column
        plug_type = str(row[7]).strip()
        if 'NEMA' in plug_type:
            current_data['NEMA Connection'] = current_data.get('NEMA Connection', '') + " " + plug_type

    if pd.notna(row[1]):  # Description column
        description = str(row[1]).strip()
        if 'shelves' in description:
            description = description.replace('35.7"', '')
            current_data['Description'] = current_data.get('Description', '') + " " + description

# Append the last collected model if there's any left
if current_data:
    data_points.append(current_data)

# Clean extra spaces in all data points
for entry in data_points:
    if 'NEMA Connection' in entry:
        entry['NEMA Connection'] = entry['NEMA Connection'].rstrip(', ')
    for key, value in entry.items():
        # Replace multiple spaces with a single space and strip leading/trailing spaces
        entry[key] = re.sub(r'\s+', ' ', value.strip())

        if key in ['Voltage', 'Watts', 'Amps']:
            entry[key] = value.split()[0]

        if key == 'External Dimensions' or key == 'Shipping Weight':
            entry[key] = value.strip()
            

# Create a new workbook using openpyxl
wb = Workbook()

# Define the color for section separator rows
separator_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
model_fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")
# Remove default sheet and add data for each model
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Iterate over each model and create a new sheet for each one
for entry in data_points:
    model_name = entry.get('Model', 'Unknown Model')
    sheet = wb.create_sheet(title=model_name)
    external_dimensions = entry.get('External Dimensions', '')

    # Set column headers: Attributes, Values, Notes
    rw = 1
    sheet.cell(rw, column=1, value="Attributes").alignment = Alignment(wrap_text=True)
    sheet.cell(rw, column=2, value="Values").alignment = Alignment(wrap_text=True)
    sheet.cell(rw, column=3, value="Notes").alignment = Alignment(wrap_text=True)
    for col in range(3):
        sheet[rw][col].fill = header_fill


    # Extract Width, Depth, Height from External Dimensions (without a separate function)
    match = re.match(r'([0-9\.]+") x ([0-9\.]+") x ([0-9\.]+")', external_dimensions)
    if match:
        width, depth, height = match.groups()  # Returns a tuple (Width, Depth, Height)
    else:
        width, depth, height = None, None, None  # Return None if no match
    # Define the sections (Attributes)

    sections = [
        ("Model", entry.get('Model', '')),
        ("External Dimensions", ""),("Width", width),("Depth", depth),("Height", height),("Shipping Info", ""),
        ("Shipping Weight (lbs)", entry.get('Shipping Weight', '')),("Certifications", ""),
        ("Electrical", ""), ("Voltage", entry.get('Voltage', '')),("Cycle", ""), ("Phase", ""), ("Amps", entry.get('Amps', '')), ("Kw", entry.get('Watts', '')),("HP (Fractions)", ""), ("NEMA Connection", entry.get('NEMA Connection', '')),
        ("Gas", ""), ("Gas Type (Natural or Propane)", ""), ("Gas Size", ""), ("Gas Total BTUs", ""), ("Gas Kw", ""),
        ("Plumbing", ""), ("Cold Water Size", ""), ("Hot Water Size", ""), ("Drain Size", ""), ("Indirect Waste Size", ""), ("Direct Waste Size", "")
    ]

    # Write sections to the sheet
    row = 2
    for section in sections:
        # Write the attribute (e.g., "Model", "Voltage", etc.)
        sheet.cell(row=row, column=1, value=section[0]).alignment = Alignment(wrap_text=True)
        sheet.cell(row=row, column=2, value=section[1]).alignment = Alignment(wrap_text=True)
        sheet.cell(row=row, column=3, value="").alignment = Alignment(wrap_text=True)

        # Apply the color to separator rows
        if section[0] == "Model":
            for col in range(3):
                sheet[row][col].fill = model_fill

        if section[0] == "External Dimensions" or section[0] == "Shipping Info" or section[0] == "Electrical" or section[0] == "Gas" or section[0] == "Plumbing":
            for col in range(3):
                sheet[row][col].fill = separator_fill
        row += 1

# Save the workbook to a file
wb.save('HATCO_SPEC_SHEET_WORKBOOK.xlsx')
print("Excel workbook 'HATCO_SPEC_SHEET_WORKBOOK.xlsx' has been created.")