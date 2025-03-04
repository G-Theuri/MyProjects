import camelot
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
import re

# Function to read and clean the PDF data
def read_and_clean_pdf(pdf_path):
    # Read the PDF and extract the first table
    df = camelot.read_pdf(pdf_path, pages='2', flavor='stream')[0].df
    df = df.dropna(how='all', axis=0).dropna(how='all', axis=1).iloc[4:]  # Clean and remove unnecessary rows
    return df

# Function to extract relevant data points from the cleaned DataFrame
def extract_data_points(df):
    data_points = []
    current_data = {}
    
    for _, row in df.iterrows():
        if 'NEMA 5-15P' in str(row[7]) and current_data:
            data_points.append(current_data)
            current_data = {}

        if pd.notna(row[0]): current_data['Model'] = str(row[0]).strip()
        if pd.notna(row[2]): current_data['External Dimensions'] = str(row[2]).strip()
        if pd.notna(row[3]): current_data['Voltage'] = str(row[3]).strip()
        if pd.notna(row[4]): current_data['Watts'] = str(row[4]).strip()
        if pd.notna(row[6]): current_data['Amps'] = str(row[6]).strip()
        if pd.notna(row[8]): current_data['Shipping Weight'] = str(row[8]).strip()
        if pd.notna(row[7]): current_data['NEMA Connection'] = str(row[7]).strip()

    if current_data: data_points.append(current_data)

    # Clean the extracted data (space handling, value extraction)
    for entry in data_points:
        entry['External Dimensions'] = re.sub(r'\s+', ' ', entry.get('External Dimensions', '').strip())
        entry['NEMA Connection'] = entry.get('NEMA Connection', '').rstrip(', ')
        for key in ['Voltage', 'Watts', 'Amps']:
            entry[key] = entry.get(key, '').split()[0] if entry.get(key, '') else '' # Keep only the first value

    return data_points

# Function to create an Excel workbook and write the extracted data
def create_workbook(data_points):
    # Create workbook and color formatting
    wb = Workbook()
    separator_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Remove default sheet and add data for each model
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Process each model's data and create a new sheet for it
    for entry in data_points:
        model_name = entry.get('Model')
        sheet = wb.create_sheet(title=model_name)

        # Extract width, depth, and height from external dimensions
        match = re.match(r'([0-9\.]+") x ([0-9\.]+") x ([0-9\.]+")', entry.get('External Dimensions', ''))
        width, depth, height = match.groups() if match else (None, None, None)

        # Sections to write to the sheet
        sections = [
            ("Model", entry.get('Model', '')),
            ("External Dimensions", ""),
            ("Width", width),
            ("Depth", depth),
            ("Height", height),
            ("Shipping Info", ""),
            ("Shipping Weight (lbs)", entry.get('Shipping Weight', '')),
            ("Certifications", ""),
            ("Electrical", ""),
            ("Voltage", entry.get('Voltage', '')),
            ("Cycle", ""),
            ("Phase", ""),
            ("Amps", entry.get('Amps', '')),
            ("Kw", entry.get('Watts', '')),
            ("HP (Fractions)", ""),
            ("NEMA Connection", entry.get('NEMA Connection', '')),
            ("Gas", ""),
            ("Gas Type (Natural or Propane)", ""),
            ("Gas Size", ""),
            ("Gas Total BTUs", ""),
            ("Gas Kw", ""),
            ("Plumbing", ""),
            ("Cold Water Size", ""),
            ("Hot Water Size", ""),
            ("Drain Size", ""),
            ("Indirect Waste Size", ""),
            ("Direct Waste Size", "")
        ]

        # Write data to sheet
        row = 1
        for section in sections:
            sheet.cell(row=row, column=1, value=section[0]).alignment = Alignment(wrap_text=True)
            sheet.cell(row=row, column=2, value=section[1]).alignment = Alignment(wrap_text=True)
            sheet.cell(row=row, column=3, value="").alignment = Alignment(wrap_text=True)

            # Apply color to separator rows
            if section[0] in ["External Dimensions", "Shipping Weight (lbs)", "Voltage", "Amps"]:
                for col in range(3):
                    sheet[row][col].fill = separator_fill
            row += 1

    # Save workbook to file
    wb.save('HATCO_SPEC_SHEET.xlsx')
    print("Excel workbook 'HATCO_SPEC_SHEET_WORKBOOK.xlsx' has been created.")

def main(pdf_path):
    df = read_and_clean_pdf(pdf_path) # Read and clean PDF
    data_points = extract_data_points(df)  # Extract data points
    create_workbook(data_points) # Create and save workbook

# Run the main function
if __name__ == "__main__":
    pdf_path = "resources/Hatco Spec Sheet.pdf"
    main(pdf_path)
